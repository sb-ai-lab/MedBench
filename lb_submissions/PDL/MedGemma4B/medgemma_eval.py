"""
Cкрипт для запуска MedGemma на всех задачах RuMedBench
"""

import json
import yaml
import time
import requests
from pathlib import Path
from typing import Dict, List, Tuple
import re
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========

def load_config(config_path: str) -> Dict:
    """Загрузка конфигурации из YAML"""
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)

def load_benchmark_data(task_name: str, split: str = "dev", max_samples: int = None) -> List[Dict]:
    """Загрузка данных бенчмарка для конкретной задачи"""
    
    # Пути к данным
    data_paths = {
        "RuMedTest": {
            "test": "data/benchmarks/rumedbench/RuMedTest/test.jsonl",
        },
        "RuMedDaNet": {
            "dev": "data/benchmarks/rumedbench/RuMedDaNet/dev.jsonl",
            "test": "data/benchmarks/rumedbench/RuMedDaNet/test.jsonl"
        },
        "RuMedNLI": {
            "dev": "data/benchmarks/rumedbench/RuMedNLI/dev.jsonl",
            "test": "data/benchmarks/rumedbench/RuMedNLI/test.jsonl"
        }
    }
    
    if task_name not in data_paths:
        raise ValueError(f"Неизвестная задача: {task_name}")
    
    if split not in data_paths[task_name]:
        available = list(data_paths[task_name].keys())
        raise ValueError(f"Сплит '{split}' недоступен для задачи '{task_name}'. Доступно: {available}")
    
    data_path = data_paths[task_name][split]
    full_path = Path(data_path)
    
    if not full_path.exists():
        # Пробуем найти в поддиректориях
        alt_path = Path("data/benchmarks/rumedbench") / task_name / f"{split}.jsonl"
        if alt_path.exists():
            full_path = alt_path
        else:
            raise FileNotFoundError(f"Файл не найден: {data_path}")
    
    data = []
    with open(full_path, 'r', encoding='utf-8') as f:
        for i, line in enumerate(f):
            if line.strip():
                item = json.loads(line)
                data.append(item)
                
                if max_samples and len(data) >= max_samples:
                    break
    
    print(f"Загружено {len(data)} примеров из {full_path}")
    return data

def create_prompt(task_name: str, item: Dict) -> str:
    """Создание промпта в зависимости от задачи"""
    
    if task_name == "RuMedTest":
        # RuMedTest: множественный выбор
        options = []
        for i in range(1, 5):
            if str(i) in item:
                options.append(item[str(i)])
        
        prompt = f"""You are a medical doctor and absolutely need to answer which one of the four statements has highest probability of being correct:
1. {item['question']} {options[0] if len(options) > 0 else ""}".
2. {item['question']} {options[1] if len(options) > 1 else ""}".
3. {item['question']} {options[2] if len(options) > 2 else ""}".
4. {item['question']} {options[3] if len(options) > 3 else ""}".
The number of the most likely correct statement is:"""
        return prompt
    
    elif task_name == "RuMedDaNet":
        # RuMedDaNet: вопросы Да/Нет
        prompt = f"""Imagine that you are a medical doctor and know everything about medicine and need to pass exam. 
The context is: {item.get('context', '')}
The question is: {item['question']}
You should answer only yes or no.
The answer is """
        return prompt
    
    elif task_name == "RuMedNLI":
        # RuMedNLI: логический вывод
        prompt = f"""You are a medical doctor and need to pass exam. You are given two statements:
The first statement is absolutely correct and should be the basis for your answer: "{item['ru_sentence1']}"
The second statement is "{item['ru_sentence2']}". 
You should answer if the second statement is 'entailment', 'contradiction', or 'neutral'.
The answer is """
        return prompt
    
    else:
        raise ValueError(f"Неизвестная задача: {task_name}")

def parse_answer(task_name: str, response: str) -> str:
    """Парсинг ответа в зависимости от задачи"""
    if not response:
        return ""
    
    cleaned = response.replace('**', '').replace('*', '').strip()
    
    if task_name == "RuMedTest":
        # Ищем цифру от 1 до 5
        patterns = [
            r'Answer[:\s]*(\d+)',
            r'answer[:\s]*(\d+)',
            r'(\d+)\.$',
            r'(\d+)\s*$',
            r'is\s*(\d+)',
            r'statement\s*(\d+)',
            r'\b(\d+)\b',
            r'\((\d+)\)',
            r'\[(\d+)\]',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, cleaned, re.IGNORECASE)
            if match:
                try:
                    answer = match.group(1)
                    if answer in ['1', '2', '3', '4', '5']:
                        return answer
                except (IndexError, AttributeError):
                    continue
        
        # Пробуем найти любую цифру
        all_numbers = re.findall(r'\d+', cleaned)
        for number in all_numbers:
            if number in ['1', '2', '3', '4', '5']:
                return number
        
        return "0"
    
    elif task_name == "RuMedDaNet":
        # Ищем да/нет или yes/no
        patterns = [
            r'\b(yes|да)\b',
            r'\b(no|нет)\b',
            r'Answer[:\s]*(yes|no|да|нет)',
            r'answer[:\s]*(yes|no|да|нет)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, cleaned, re.IGNORECASE)
            if match:
                answer = match.group(1).lower()
                # Преобразуем в русский
                if answer in ['yes', 'да']:
                    return 'да'
                elif answer in ['no', 'нет']:
                    return 'нет'
        
        # Пробуем по первым буквам
        if cleaned.lower().startswith('да') or cleaned.lower().startswith('yes'):
            return 'да'
        elif cleaned.lower().startswith('нет') or cleaned.lower().startswith('no'):
            return 'нет'
        
        return 'нет'  # По умолчанию
    
    elif task_name == "RuMedNLI":
        # Ищем entailment, contradiction, neutral
        patterns = [
            r'\b(entailment|contradiction|neutral)\b',
            r'Answer[:\s]*(entailment|contradiction|neutral)',
            r'answer[:\s]*(entailment|contradiction|neutral)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, cleaned, re.IGNORECASE)
            if match:
                return match.group(1).lower()
        
        # Пробуем по первым буквам
        if 'entail' in cleaned.lower():
            return 'entailment'
        elif 'contrad' in cleaned.lower():
            return 'contradiction'
        elif 'neutral' in cleaned.lower():
            return 'neutral'
        
        return 'neutral'  # По умолчанию
    
    return ""

def get_gold_answer(task_name: str, item: Dict) -> str:
    """Получение правильного ответа"""
    if task_name == "RuMedTest":
        # Для RuMedTest нет ответов в test данных
        for key in ['answer', 'Answer', 'correct_answer', 'label', 'gold']:
            if key in item:
                return str(item[key])
        return ""  # Пусто для test данных
    
    elif task_name == "RuMedDaNet":
        return item.get('answer', '')
    
    elif task_name == "RuMedNLI":
        return item.get('gold_label', '')
    
    return ""

def call_model(prompt: str, model_config: Dict, max_retries: int = 3) -> str:
    """Вызов модели через API"""
    api_config = model_config['model']['api']
    params = model_config['model']['parameters']
    
    payload = {
        "model": api_config['model_name'],
        "messages": [
            {"role": "user", "content": prompt}
        ],
        "temperature": params['temperature'],
        "max_tokens": params['max_tokens'],
        "top_p": params['top_p']
    }
    
    for attempt in range(max_retries):
        try:
            response = requests.post(
                url=api_config['base_url'] + api_config['endpoint'],
                headers=api_config['headers'],
                json=payload,
                timeout=api_config['timeout']
            )
            
            if response.status_code == 200:
                result = response.json()
                content = result['choices'][0]['message']['content']
                return content.strip()
            else:
                print(f"Ошибка API (попытка {attempt + 1}): {response.status_code}")
                
        except Exception as e:
            print(f"Ошибка подключения (попытка {attempt + 1}): {str(e)}")
        
        if attempt < max_retries - 1:
            wait_time = 5 * (attempt + 1)
            time.sleep(wait_time)
    
    return ""

def calculate_metrics(task_name: str, results: List[Dict]) -> Dict:
    """Расчет метрик"""
    if not results:
        return {"task": task_name, "total": 0, "has_gold_answers": False}
    
    total = len(results)
    
    # Фильтруем только те примеры, где есть правильные ответы
    results_with_gold = [r for r in results if r.get('gold') and r['gold'] != ""]
    has_gold_answers = len(results_with_gold) > 0
    
    if not has_gold_answers:
        return {
            "task": task_name,
            "total": total,
            "has_gold_answers": False,
            "note": "В данных нет правильных ответов для расчёта метрик"
        }
    
    correct = sum(1 for r in results_with_gold if r['predicted'] == r['gold'])
    accuracy = correct / len(results_with_gold) * 100 if len(results_with_gold) > 0 else 0
    
    return {
        "task": task_name,
        "total": total,
        "with_gold_answers": len(results_with_gold),
        "correct": correct,
        "incorrect": len(results_with_gold) - correct,
        "accuracy": accuracy,
        "accuracy_percent": f"{accuracy:.2f}%",
        "has_gold_answers": True
    }

def save_results(results: List[Dict], metrics: Dict, model_name: str, task_name: str, split: str = "dev"):
    """Сохранение результатов"""
    output_dir = Path("results") / model_name / task_name
    output_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    # Сохраняем детальные результаты
    detailed_file = output_dir / f"detailed_{split}_{timestamp}.json"
    with open(detailed_file, 'w', encoding='utf-8') as f:
        json.dump({
            "task": task_name,
            "split": split,
            "model": model_name,
            "timestamp": timestamp,
            "metrics": metrics,
            "results": results
        }, f, indent=2, ensure_ascii=False)
    
    # Сохраняем предсказания
    predictions_file = output_dir / f"predictions_{split}_{timestamp}.jsonl"
    with open(predictions_file, 'w', encoding='utf-8') as f:
        for result in results:
            if task_name == "RuMedTest":
                submit_item = {"idx": result["idx"], "answer": str(result["predicted"])}
            elif task_name == "RuMedDaNet":
                submit_item = {"pairID": result["idx"], "answer": result["predicted"]}
            elif task_name == "RuMedNLI":
                submit_item = {"pairID": result["idx"], "gold_label": result["predicted"]}
            else:
                submit_item = {"id": result["idx"], "answer": result["predicted"]}
            
            f.write(json.dumps(submit_item, ensure_ascii=False) + '\n')
    
    print(f"  Детальные результаты: {detailed_file.name}")
    print(f"  Предсказания: {predictions_file.name}")
    
    return output_dir

# ========== ОСНОВНЫЕ ФУНКЦИИ ==========

def run_task(task_name: str, model_config: Dict, max_samples: int = 10, split: str = "dev") -> Tuple[List[Dict], Dict]:
    """Запуск одной задачи"""
    print(f"\n{'='*60}")
    print(f"Запуск задачи: {task_name} (сплит: {split})")
    print(f"{'='*60}")
    
    # Загружаем данные
    try:
        data = load_benchmark_data(task_name, split=split, max_samples=max_samples)
    except FileNotFoundError as e:
        print(f"Ошибка: {str(e)}")
        return [], {"error": str(e)}
    
    if not data:
        print("Нет данных для обработки")
        return [], {}
    
    results = []
    
    for i, item in enumerate(data, 1):
        print(f"\n[{i}/{len(data)}] Пример {i}")
        
        # Для отладки выводим информацию
        if task_name == "RuMedTest":
            question_preview = item.get('question', '')[:80]
            print(f"Вопрос: {question_preview}...")
        elif task_name == "RuMedDaNet":
            question_preview = item.get('question', '')[:80]
            print(f"Вопрос: {question_preview}...")
        elif task_name == "RuMedNLI":
            s1_preview = item.get('ru_sentence1', '')[:60]
            s2_preview = item.get('ru_sentence2', '')[:60]
            print(f"Утверждение 1: {s1_preview}...")
            print(f"Утверждение 2: {s2_preview}...")
        
        # Создаем промпт
        prompt = create_prompt(task_name, item)
        
        # Вызываем модель
        response = call_model(prompt, model_config)
        
        # Парсим ответ
        predicted = parse_answer(task_name, response)
        gold = get_gold_answer(task_name, item)
        
        # Сохраняем результат
        result = {
            "idx": item.get('idx', item.get('pairID', str(i))),
            "prompt_preview": prompt[:100] + "..." if len(prompt) > 100 else prompt,
            "response_preview": response[:100] + "..." if len(response) > 100 else response,
            "predicted": predicted,
            "gold": gold,
            "has_gold_answer": bool(gold and gold != ""),
            "correct": predicted == gold if gold else None
        }
        
        results.append(result)
        
        print(f"Ответ модели: '{predicted}'")
        if gold:
            print(f"Правильный ответ: '{gold}'")
            print(f"Верно: {predicted == gold}")
        else:
            print("Нет правильного ответа в данных")
        
        # Пауза между запросами
        if i < len(data):
            time.sleep(2)
    
    # Расчет метрик
    metrics = calculate_metrics(task_name, results)
    
    print(f"\nРезультаты {task_name}:")
    for key, value in metrics.items():
        print(f"  {key}: {value}")
    
    return results, metrics

def create_excel_summary(model_name: str, all_metrics: Dict, tasks_config: List[Dict]):
    """Создание Excel отчета с итоговой статистикой"""
    
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    summary_dir = Path("results") / model_name
    summary_dir.mkdir(parents=True, exist_ok=True)
    
    excel_file = summary_dir / f"summary_{timestamp}.xlsx"
    
    # Создаем DataFrame с метриками
    data = []
    for task_config in tasks_config:
        task = task_config["name"]
        split = task_config["split"]
        
        metrics = all_metrics.get(task, {})
        
        row = {
            "Task": task,
            "Split": split,
            "Total Samples": metrics.get("total", 0),
            "With Gold Answers": metrics.get("with_gold_answers", metrics.get("total", 0)),
            "Correct": metrics.get("correct", 0),
            "Incorrect": metrics.get("incorrect", 0),
            "Accuracy (%)": metrics.get("accuracy", 0),
            "Accuracy (display)": metrics.get("accuracy_percent", "0.00%"),
            "Has Gold": "✓" if metrics.get("has_gold_answers", False) else "✗"
        }
        data.append(row)
    
    # Основной DataFrame
    df_main = pd.DataFrame(data)
    
    # Итоговая строка
    total_with_gold = sum(m.get("with_gold_answers", m.get("total", 0)) 
                         for m in all_metrics.values() if m.get("has_gold_answers", False))
    total_correct = sum(m.get("correct", 0) for m in all_metrics.values())
    
    if total_with_gold > 0:
        overall_accuracy = (total_correct / total_with_gold) * 100
    else:
        overall_accuracy = 0
    
    summary_row = {
        "Task": "OVERALL",
        "Split": "",
        "Total Samples": df_main["Total Samples"].sum(),
        "With Gold Answers": total_with_gold,
        "Correct": total_correct,
        "Incorrect": total_with_gold - total_correct,
        "Accuracy (%)": overall_accuracy,
        "Accuracy (display)": f"{overall_accuracy:.2f}%",
        "Has Gold": "✓" if total_with_gold > 0 else "✗"
    }
    
    df_summary = pd.DataFrame([summary_row])
    df_final = pd.concat([df_main, df_summary], ignore_index=True)
    
    # Создаем Excel файл с оформлением
    wb = Workbook()
    ws = wb.active
    ws.title = "Results Summary"
    
    # Записываем данные
    for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Настройка стилей
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    # Данные
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            
            # Центрируем числа
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="center")
            
            # Заливка для итоговой строки
            if cell.row == ws.max_row:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                cell.font = Font(bold=True)
            
            # Цвета для accuracy
            if cell.column == 7:  # Accuracy (%)
                if isinstance(cell.value, (int, float)):
                    if cell.value >= 80:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # зеленый
                    elif cell.value >= 60:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # желтый
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # красный
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Лист с деталями по задачам
    for task, metrics in all_metrics.items():
        if "results" in metrics:
            ws_details = wb.create_sheet(title=f"{task[:25]}")  # ограничиваем имя листа
            results = metrics["results"]
            
            if results:
                # Конвертируем результаты в DataFrame
                df_details = pd.DataFrame(results)
                
                # Оставляем только важные колонки
                important_cols = ["idx", "predicted", "gold", "correct", "response_preview"]
                available_cols = [col for col in important_cols if col in df_details.columns]
                
                if available_cols:
                    df_details = df_details[available_cols]
                    
                    # Записываем данные
                    for r_idx, row in enumerate(dataframe_to_rows(df_details, index=False, header=True), 1):
                        for c_idx, value in enumerate(row, 1):
                            ws_details.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Стили для заголовков
                    for cell in ws_details[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Лист с метаинформацией
    ws_meta = wb.create_sheet(title="Metadata")
    meta_data = [
        ["Model", model_name],
        ["Evaluation Date", time.strftime("%Y-%m-%d %H:%M:%S")],
        ["Evaluation Type", "Zero-shot"],
        ["API Endpoint", "LM Studio"],
        ["", ""],
        ["Tasks Configuration", ""]
    ]
    
    for task_config in tasks_config:
        meta_data.append([f"{task_config['name']} split", task_config['split']])
    
    for row_idx, (key, value) in enumerate(meta_data, 1):
        ws_meta.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws_meta.cell(row=row_idx, column=2, value=value)
    
    # Сохраняем
    wb.save(excel_file)
    
    print(f"\n✅ Excel отчет сохранен: {excel_file}")
    return excel_file
def main():
    """Основная функция"""
    print("=" * 60)
    print("Запуск MedGemma 4B на всех задачах RuMedBench")
    print("=" * 60)
    
    # Конфиги
    model_config_path = "configs/models/medgemma_4b_it.yaml"
    
    if not Path(model_config_path).exists():
        print(f"Ошибка: файл не найден - {model_config_path}")
        return
    
    # Загружаем конфиг модели
    print("Загрузка конфигурации модели...")
    model_config = load_config(model_config_path)
    model_name = model_config['model']['name']
    print(f"Модель: {model_name}")
    
    # Список задач для запуска с разными сплитами
    tasks_config = [
        {"name": "RuMedTest", "split": "test"},  # Только test данные
        {"name": "RuMedDaNet", "split": "test"},  # dev данные (есть ответы)
        {"name": "RuMedNLI", "split": "test"}     # dev данные (есть ответы)
    ]
    
    # Параметры запуска
    max_samples_per_task = None  # Для теста
    total_start_time = time.time()
    all_metrics = {}
    
    for task_config in tasks_config:
        task = task_config["name"]
        split = task_config["split"]
        
        try:
            # Запускаем задачу
            task_start_time = time.time()
            results, metrics = run_task(task, model_config, 
                                      max_samples=max_samples_per_task,
                                      split=split)
            task_time = time.time() - task_start_time
            
            if results:
                # Сохраняем результаты
                output_dir = save_results(results, metrics, model_name, task, split)
                all_metrics[task] = metrics
                
                print(f"\nЗадача '{task}' завершена за {task_time:.2f} секунд")
                print(f"Результаты сохранены в: {output_dir}")
            
            # Пауза между задачами
            if task_config != tasks_config[-1]:
                print(f"\nПауза 5 секунд перед следующей задачей...")
                time.sleep(5)
                
        except Exception as e:
            print(f"\nОшибка при выполнении задачи '{task}': {str(e)}")
            import traceback
            traceback.print_exc()
            continue
    
    # Итоговый отчет
    total_time = time.time() - total_start_time
    print(f"\n{'='*60}")
    print("ИТОГОВЫЙ ОТЧЕТ")
    print(f"{'='*60}")
    print(f"Общее время выполнения: {total_time:.2f} секунд")
    print(f"Количество задач: {len(all_metrics)}")
    
    if all_metrics:
        print(f"\nМетрики по задачам:")
        for task, metrics in all_metrics.items():
            print(f"\n{task}:")
            if metrics.get('has_gold_answers', False):
                for key, value in metrics.items():
                    if key not in ['task', 'has_gold_answers']:
                        print(f"  {key}: {value}")
            else:
                print(f"  {metrics.get('note', 'Нет метрик (test данные без ответов)')}")
        
        # Сохраняем итоговый отчет
        summary_dir = Path("results") / model_name
        summary_dir.mkdir(parents=True, exist_ok=True)
        summary_file = summary_dir / f"summary_{time.strftime('%Y%m%d_%H%M%S')}.json"
        
        with open(summary_file, 'w', encoding='utf-8') as f:
            json.dump({
                "model": model_name,
                "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                "total_time_seconds": total_time,
                "tasks": all_metrics,
                "parameters": {
                    "max_samples_per_task": max_samples_per_task
                }
            }, f, indent=2, ensure_ascii=False)
        
        print(f"\nИтоговый отчет сохранен: {summary_file}")
        excel_file = create_excel_summary(model_name, all_metrics, tasks_config)
        print(f"Excel отчет сохранен: {excel_file}")
    
    print(f"\nЗавершено!")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nПрервано пользователем")
        sys.exit(0)
    except Exception as e:
        print(f"\nОшибка: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)